import openpyxl

wb = openpyxl.load_workbook('csdl.xlsx')
print("All sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames[:15]:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print('='*60)
    ws = wb[sheet_name]
    max_row = ws.max_row
    print(f"Rows: {max_row}")
    for i, row in enumerate(ws.iter_rows(max_col=10, max_row=min(6, max_row+1), values_only=True)):
        print(row)
